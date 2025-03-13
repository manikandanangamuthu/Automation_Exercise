
#File--> workbook-->Sheets-->Rows--Cells


import openpyxl  # This module required for work with excel file

from openpyxl.styles import PatternFill  # This is for coloring

#path="C://Users//HP//Desktop//selenium_Excel//Excel_functions.xlsx"

def getRowsCount(file_path,sheetName):
    workbook=openpyxl.load_workbook(file_path)
    sheet=workbook[sheetName]
    return(sheet.max_row)


def getColumnCount(file_path,sheetName):
    workbook=openpyxl.load_workbook(file_path)
    sheet=workbook[sheetName]
    return (sheet.max_column)

def readData(file_path,sheetName,row_no,column_no):
    workbook=openpyxl.load_workbook(file_path)
    sheet=workbook[sheetName]
    return sheet.cell(row_no,column_no).value

def writeData(file_path,sheetName,row_no,column_no,data):
    workbook=openpyxl.load_workbook(file_path)
    sheet=workbook[sheetName]
    sheet.cell(row_no,column_no).value=data
    workbook.save(file_path)

def fillGreencolor(file_path,sheetName,row_no,column_no):
    workbook=openpyxl.load_workbook(file_path)
    sheet=workbook[sheetName]
    greenFill=PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')
    sheet.cell(row_no,column_no).fill=greenFill
    workbook.svae(file_path)



def fillRedcolor(file_path,sheetName,row_no,column_no):
    workbook=openpyxl.load_workbook(file_path)
    sheet=workbook[sheetName]
    redFill=PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    sheet.cell(row_no,column_no).fill=redFill
    workbook.svae(file_path)